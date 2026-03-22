"""
Risk Dashboard — Portefeuille Complet
======================================
Agrège toutes les positions :
  - Momentum bot (ib_paper_trading.py, clientId 1)
  - PEAD (arb1, clientId 2)
  - Spinoff (arb2, clientId 3)
  - Index Rebal (arb3, clientId 4)
  - Convertibles equity hedge (arb4, clientId 5)

Calculs en temps réel :
  - P&L par position et global
  - VaR 95% / 99% (historique + paramétrique)
  - CVaR (Expected Shortfall)
  - Beta portefeuille vs SPY
  - Sharpe / Sortino (rolling 30j)
  - Max Drawdown courant
  - Stress tests : -10%, -20%, -30% marché
  - Concentration / corrélation

Export Excel à la demande : tapez 'x' dans le terminal
Rafraîchissement : toutes les 60s (configurable)
"""

import sys
import os
import threading
import time
import numpy as np
import pandas as pd
from ib_insync import IB, Stock, util
from datetime import datetime, timedelta
import logging
import requests
from alerts import alert_manager

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.WARNING,
    handlers=[logging.FileHandler('risk_dashboard.log', encoding='utf-8')]
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
HOST      = '127.0.0.1'
PORT         = 4002
CLIENT_ID  = 20         # clientId dédié au risk dashboard (dashboard_server utilise 9)
CAPITAL    = 1_090_000
REFRESH_S  = 60         # rafraîchissement en secondes

# CSV des stratégies arb
PEAD_CSV        = 'pead_positions.csv'
SPINOFF_CSV     = 'spinoff_positions.csv'
INDEX_REBAL_CSV = 'index_rebal_positions.csv'
CONV_CSV        = 'convertibles_positions.csv'

# Historique prix pour VaR (en mémoire)
price_history: dict[str, list[float]] = {}
pnl_history:   list[float] = []

# ── Connexion IB ─────────────────────────────────────────────────────────────

def connect() -> IB:
    ib = IB()
    ib.connect(HOST, PORT, clientId=CLIENT_ID)
    return ib


def get_price(ib: IB, ticker: str) -> float | None:
    contract = Stock(ticker, 'SMART', 'USD')
    try:
        qualified = ib.qualifyContracts(contract)
        if not qualified:
            return None
        ib.reqMarketDataType(3)
        ticker_data = ib.reqMktData(qualified[0], '', False, False)
        ib.sleep(3)
        price = ticker_data.last or ticker_data.close or ticker_data.bid
        ib.cancelMktData(qualified[0])
        return float(price) if price and price > 0 else None
    except Exception:
        return None


def get_spy_return(days: int = 30) -> float:
    """Retour SPY sur N jours via Yahoo Finance (fallback si IB indispo)."""
    try:
        end   = datetime.now()
        start = end - timedelta(days=days + 5)
        url   = (f"https://query1.finance.yahoo.com/v8/finance/chart/SPY"
                 f"?interval=1d&period1={int(start.timestamp())}&period2={int(end.timestamp())}")
        r = requests.get(url, timeout=5, headers={'User-Agent': 'Mozilla/5.0'})
        closes = r.json()['chart']['result'][0]['indicators']['quote'][0]['close']
        closes = [c for c in closes if c]
        if len(closes) >= 2:
            return (closes[-1] - closes[0]) / closes[0]
    except Exception:
        pass
    return 0.0

# ── Lecture positions IB live ─────────────────────────────────────────────────

def get_ib_positions(ib: IB) -> pd.DataFrame:
    rows = []
    for pos in ib.positions():
        if pos.position == 0:
            continue
        rows.append({
            'ticker':    pos.contract.symbol,
            'shares':    pos.position,
            'avg_cost':  pos.avgCost,
            'direction': 'LONG' if pos.position > 0 else 'SHORT',
        })
    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['ticker', 'shares', 'avg_cost', 'direction'])


def get_portfolio_items(ib: IB) -> dict:
    """Retourne {ticker: market_price} depuis le portfolio IB."""
    prices = {}
    for item in ib.portfolio():
        prices[item.contract.symbol] = {
            'market_price': item.marketPrice,
            'market_value': item.marketValue,
            'unrealized_pnl': item.unrealizedPNL,
        }
    return prices

# ── Lecture CSV arb ───────────────────────────────────────────────────────────

def load_arb_data() -> pd.DataFrame:
    """
    Charge tous les CSV arb et retourne un DataFrame unifié.
    Colonnes garanties : ticker, strategy, entry_price, entry_date
    Colonnes optionnelles selon source : surprise, spinoff_name, index, days_to_rebal,
                                         bond_price, theo_value, mispricing, delta
    Source TWS = vérité sur shares/avg_cost/mkt_price (prix live).
    Source CSV = enrichissement : stratégie, prix d'entrée original, date, métadonnées.
    """
    frames = []

    # ── PEAD ──────────────────────────────────────────────────────────────────
    if os.path.exists(PEAD_CSV):
        try:
            df = pd.read_csv(PEAD_CSV)
            if not df.empty:
                df = df.rename(columns={'entry_price': 'csv_entry_price'})
                df['strategy'] = 'PEAD'
                # Calcul jours détenus
                df['entry_date'] = pd.to_datetime(df['entry_date'], errors='coerce')
                df['days_held'] = (datetime.now() - df['entry_date']).dt.days.fillna(0).astype(int)
                frames.append(df[['ticker','strategy','csv_entry_price','entry_date',
                                   'days_held','surprise']])
        except Exception as e:
            log.warning(f"PEAD CSV error: {e}")

    # ── Spinoff ───────────────────────────────────────────────────────────────
    if os.path.exists(SPINOFF_CSV):
        try:
            df = pd.read_csv(SPINOFF_CSV)
            if not df.empty:
                df = df.rename(columns={'entry_price': 'csv_entry_price'})
                df['strategy'] = 'Spinoff'
                df['entry_date'] = pd.to_datetime(df['entry_date'], errors='coerce')
                df['days_held'] = (datetime.now() - df['entry_date']).dt.days.fillna(0).astype(int)
                frames.append(df[['ticker','strategy','csv_entry_price','entry_date','days_held']])
        except Exception as e:
            log.warning(f"Spinoff CSV error: {e}")

    # ── Index Rebal ───────────────────────────────────────────────────────────
    if os.path.exists(INDEX_REBAL_CSV):
        try:
            df = pd.read_csv(INDEX_REBAL_CSV)
            if not df.empty:
                df = df.rename(columns={'entry_price': 'csv_entry_price'})
                df['strategy'] = 'IndexRebal'
                df['entry_date'] = pd.to_datetime(df['entry_date'], errors='coerce')
                df['days_held'] = (datetime.now() - df['entry_date']).dt.days.fillna(0).astype(int)
                frames.append(df[['ticker','strategy','csv_entry_price','entry_date','days_held']])
        except Exception as e:
            log.warning(f"IndexRebal CSV error: {e}")

    # ── Convertibles ──────────────────────────────────────────────────────────
    if os.path.exists(CONV_CSV):
        try:
            df = pd.read_csv(CONV_CSV)
            if not df.empty:
                df['strategy'] = 'Convertibles'
                df['csv_entry_price'] = df['bond_price']  # prix obligataire
                df['entry_date'] = pd.to_datetime(df['entry_date'], errors='coerce')
                df['days_held'] = (datetime.now() - df['entry_date']).dt.days.fillna(0).astype(int)
                keep = ['ticker','strategy','csv_entry_price','entry_date',
                        'days_held','bond_price','theo_value','mispricing','delta','hedge_shares']
                frames.append(df[[c for c in keep if c in df.columns]])
        except Exception as e:
            log.warning(f"Convertibles CSV error: {e}")

    if not frames:
        return pd.DataFrame(columns=['ticker','strategy','csv_entry_price','entry_date','days_held'])

    combined = pd.concat(frames, ignore_index=True)
    # En cas de doublon ticker (même ticker dans 2 strats), garder les deux
    return combined


def load_arb_meta() -> dict:
    """Rétrocompat : retourne ticker -> strategy pour l'affichage rapide."""
    df = load_arb_data()
    if df.empty:
        return {}
    return dict(zip(df['ticker'], df['strategy']))

# ── Calculs de risque ─────────────────────────────────────────────────────────

def compute_returns(prices: list[float]) -> np.ndarray:
    arr = np.array(prices)
    if len(arr) < 2:
        return np.array([0.0])
    return np.diff(arr) / arr[:-1]


def var_historical(returns: np.ndarray, confidence: float = 0.95) -> float:
    if len(returns) < 5:
        return 0.0
    return float(-np.percentile(returns, (1 - confidence) * 100))


def var_parametric(returns: np.ndarray, confidence: float = 0.95) -> float:
    if len(returns) < 5:
        return 0.0
    mu, sigma = returns.mean(), returns.std()
    from scipy.stats import norm
    return float(-(mu + sigma * norm.ppf(1 - confidence)))


def cvar(returns: np.ndarray, confidence: float = 0.95) -> float:
    """Expected Shortfall — moyenne des pertes au-delà de la VaR."""
    if len(returns) < 5:
        return 0.0
    threshold = np.percentile(returns, (1 - confidence) * 100)
    tail = returns[returns <= threshold]
    return float(-tail.mean()) if len(tail) > 0 else 0.0


def sharpe(returns: np.ndarray, risk_free_daily: float = 0.045 / 252) -> float:
    if len(returns) < 5 or returns.std() == 0:
        return 0.0
    return float((returns.mean() - risk_free_daily) / returns.std() * np.sqrt(252))


def sortino(returns: np.ndarray, risk_free_daily: float = 0.045 / 252) -> float:
    downside = returns[returns < 0]
    if len(downside) < 2 or downside.std() == 0:
        return 0.0
    return float((returns.mean() - risk_free_daily) / downside.std() * np.sqrt(252))


def max_drawdown(nav_series: list[float]) -> float:
    if len(nav_series) < 2:
        return 0.0
    arr   = np.array(nav_series)
    peaks = np.maximum.accumulate(arr)
    dds   = (arr - peaks) / peaks
    return float(dds.min())


def beta_vs_spy(port_returns: np.ndarray, spy_returns: np.ndarray) -> float:
    n = min(len(port_returns), len(spy_returns))
    if n < 5:
        return 1.0
    p, s = port_returns[-n:], spy_returns[-n:]
    cov   = np.cov(p, s)[0, 1]
    var_s = np.var(s)
    return float(cov / var_s) if var_s > 0 else 1.0


def stress_test(total_long: float, total_short: float,
                scenarios: list[float] = [-0.10, -0.20, -0.30, 0.10]) -> dict:
    """
    Simule P&L pour différents chocs marché.
    Long positions subissent le choc, shorts profitent du choc baissier.
    """
    results = {}
    for shock in scenarios:
        pnl_long  =  total_long  * shock
        pnl_short = -total_short * shock   # short profite si marché baisse
        results[f"{shock:+.0%}"] = pnl_long + pnl_short
    return results

# ── Construction du snapshot portefeuille ─────────────────────────────────────

def build_snapshot(ib: IB) -> dict:
    # Source 1 : TWS (prix live, shares réelles, avg_cost réel)
    positions = get_ib_positions(ib)
    portfolio = get_portfolio_items(ib)

    # Source 2 : CSV arb (stratégie, prix entrée, date, métadonnées)
    arb_df   = load_arb_data()
    arb_meta = load_arb_meta()

    csv_by_ticker: dict = {}
    for _, row in arb_df.iterrows():
        t = row['ticker']
        if t not in csv_by_ticker:
            csv_by_ticker[t] = row.to_dict()

    if positions.empty:
        return {}

    rows = []
    total_market_value = 0.0
    total_cost         = 0.0

    for _, pos in positions.iterrows():
        t = pos['ticker']

        # Prix live depuis TWS (source primaire)
        info  = portfolio.get(t, {})
        mkt_p = info.get('market_price', pos['avg_cost'])
        mkt_v = info.get('market_value', mkt_p * pos['shares'])
        upnl  = info.get('unrealized_pnl', (mkt_p - pos['avg_cost']) * pos['shares'])
        cost  = pos['avg_cost'] * abs(pos['shares'])

        # Stratégie : CSV prioritaire
        csv_row  = csv_by_ticker.get(t, {})
        strategy = csv_row.get('strategy') or arb_meta.get(t, 'Momentum')
        if t == 'MSTR' and pos['shares'] < 0:
            strategy = 'Convertibles'

        # Enrichissement CSV
        entry_date = csv_row.get('entry_date', None)
        days_held  = csv_row.get('days_held',  None)
        csv_entry  = csv_row.get('csv_entry_price', None)
        surprise   = csv_row.get('surprise',   None)
        mispricing = csv_row.get('mispricing', None)

        # Détection divergence TWS vs CSV
        source_note = 'TWS+CSV' if csv_row else 'TWS only'
        if csv_entry and abs(float(csv_entry) - pos['avg_cost']) / max(pos['avg_cost'], 0.01) > 0.02:
            source_note = 'TWS≠CSV'

        weight = abs(mkt_v) / CAPITAL * 100
        meta_str = (f"Surprise:{float(surprise):.1f}%" if surprise is not None and str(surprise) != 'nan'
                    else f"Misprice:{float(mispricing):.1%}" if mispricing is not None and str(mispricing) != 'nan'
                    else '')

        rows.append({
            'Ticker':             t,
            'Stratégie':          strategy,
            'Direction':          pos['direction'],
            'Shares (TWS)':       int(pos['shares']),
            'Avg Cost (TWS)':     round(pos['avg_cost'], 3),
            'Entry (CSV)':        round(float(csv_entry), 3) if csv_entry else None,
            'Mkt Price':          round(mkt_p, 3),
            'Mkt Value ($)':      round(mkt_v, 2),
            'Cost Basis ($)':     round(cost, 2),
            'Unreal. P&L ($)':    round(upnl, 2),
            'Return (%)':         round(upnl / cost * 100, 2) if cost > 0 else 0.0,
            'Weight (%)':         round(weight, 2),
            'Days Held':          int(days_held) if days_held is not None else None,
            'Entry Date':         str(entry_date)[:10] if entry_date else None,
            'Metadata':           meta_str,
            'Source':             source_note,
        })

        total_market_value += mkt_v
        total_cost         += cost if pos['direction'] == 'LONG' else -cost

        # Historique prix pour VaR (TWS = vérité)
        if t not in price_history:
            price_history[t] = []
        price_history[t].append(mkt_p)

    df = pd.DataFrame(rows)

    # ── Métriques agrégées ────────────────────────────────────────────────────
    total_upnl   = df['Unreal. P&L ($)'].sum()
    total_long_v = df[df['Direction'] == 'LONG']['Mkt Value ($)'].sum()
    total_short_v= abs(df[df['Direction'] == 'SHORT']['Mkt Value ($)'].sum())
    gross_exp    = total_long_v + total_short_v
    net_exp      = total_long_v - total_short_v
    cash         = CAPITAL - total_long_v + total_short_v

    # P&L portfolio returns (approximation)
    nav_current = CAPITAL + total_upnl
    pnl_history.append(nav_current)

    # ── Alerte drawdown ──────────────────────────────────────────────────
    nav_peak = max(pnl_history) if pnl_history else CAPITAL
    alert_manager.check_drawdown(nav_current, nav_peak)

    port_returns = compute_returns(pnl_history)
    spy_ret      = get_spy_return(30)

    metrics = {
        'timestamp':        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'nav':              round(CAPITAL + total_upnl, 2),
        'total_upnl':       round(total_upnl, 2),
        'total_upnl_pct':   round(total_upnl / CAPITAL * 100, 2),
        'cash':             round(cash, 2),
        'gross_exposure':   round(gross_exp, 2),
        'net_exposure':     round(net_exp, 2),
        'n_positions':      len(df),
        'n_long':           len(df[df['Direction'] == 'LONG']),
        'n_short':          len(df[df['Direction'] == 'SHORT']),

        # VaR
        'var_95_hist':      round(var_historical(port_returns, 0.95) * (CAPITAL + total_upnl), 2),
        'var_99_hist':      round(var_historical(port_returns, 0.99) * (CAPITAL + total_upnl), 2),
        'var_95_param':     round(var_parametric(port_returns, 0.95) * (CAPITAL + total_upnl), 2),
        'cvar_95':          round(cvar(port_returns, 0.95) * (CAPITAL + total_upnl), 2),

        # Ratios
        'sharpe':           round(sharpe(port_returns), 3),
        'sortino':          round(sortino(port_returns), 3),
        'max_drawdown_pct': round(max_drawdown(pnl_history) * 100, 2),
        'beta':             round(beta_vs_spy(port_returns, np.array([spy_ret / 30] * len(port_returns))), 3),

        # Stress
        'stress':           stress_test(total_long_v, total_short_v),
    }

    # Concentration par stratégie
    strat_alloc = df.groupby('Stratégie')['Mkt Value ($)'].sum().to_dict()

    return {
        'positions': df,
        'metrics':   metrics,
        'strat_alloc': strat_alloc,
    }

# ── Affichage console ─────────────────────────────────────────────────────────

def print_dashboard(snap: dict):
    if not snap:
        print("⚠️  Aucune position détectée.")
        return

    m   = snap['metrics']
    df  = snap['positions']
    sa  = snap['strat_alloc']

    os.system('cls' if os.name == 'nt' else 'clear')

    ts = m['timestamp']
    nav = m['nav']
    upnl = m['total_upnl']
    upnl_pct = m['total_upnl_pct']
    sign = '+' if upnl >= 0 else ''

    print("=" * 72)
    print(f"  📊  RISK DASHBOARD — {ts}")
    print("=" * 72)

    # NAV
    print(f"\n  NAV            ${nav:>10,.2f}   ({sign}{upnl_pct:.2f}%  /  {sign}${upnl:,.2f})")
    print(f"  Cash           ${m['cash']:>10,.2f}")
    print(f"  Gross Exposure ${m['gross_exposure']:>10,.2f}   ({m['gross_exposure']/CAPITAL*100:.0f}% du capital)")
    print(f"  Net Exposure   ${m['net_exposure']:>10,.2f}")
    print(f"  Positions      {m['n_positions']}  ({m['n_long']} long / {m['n_short']} short)")

    # Allocation par stratégie
    print(f"\n  {'─'*68}")
    print("  ALLOCATION PAR STRATÉGIE")
    print(f"  {'─'*68}")
    for strat, val in sorted(sa.items(), key=lambda x: -abs(x[1])):
        pct = abs(val) / CAPITAL * 100
        bar = '█' * int(pct / 2)
        print(f"  {strat:<18} ${val:>8,.0f}  ({pct:5.1f}%)  {bar}")

    # Positions — dual source TWS + CSV
    print(f"\n  {'─'*92}")
    print("  POSITIONS  [TWS = prix live | CSV = métadonnées stratégie]")
    print(f"  {'─'*92}")
    print(f"  {'Ticker':<7} {'Dir':<6} {'Sh':>6} {'AvgCost(TWS)':>12} {'Entry(CSV)':>10} "
          f"{'MktPx':>8} {'P&L$':>9} {'Ret%':>6} {'Days':>5} {'Strat':<14} {'Src'}")
    print(f"  {'─'*92}")
    for _, r in df.sort_values('Mkt Value ($)', ascending=False).iterrows():
        sign_p = '+' if r['Unreal. P&L ($)'] >= 0 else ''
        entry  = f"{r['Entry (CSV)']:.3f}" if r['Entry (CSV)'] is not None else '    —   '
        days   = str(int(r['Days Held'])) if r['Days Held'] is not None else '—'
        meta   = f"  [{r['Metadata']}]" if r['Metadata'] else ''
        print(f"  {r['Ticker']:<7} {r['Direction']:<6} {int(r['Shares (TWS)']):>6} "
              f"{r['Avg Cost (TWS)']:>12.3f} {entry:>10} "
              f"{r['Mkt Price']:>8.3f} "
              f"{sign_p}{r['Unreal. P&L ($)']:>8.2f} "
              f"{sign_p}{r['Return (%)']:>5.2f}% "
              f"  {days:>4}  {r['Stratégie']:<14} {r['Source']}{meta}")

    # Risque
    print(f"\n  {'─'*68}")
    print("  MÉTRIQUES DE RISQUE")
    print(f"  {'─'*68}")
    print(f"  VaR 95% (hist)   -${m['var_95_hist']:,.0f}   |  VaR 99% (hist)  -${m['var_99_hist']:,.0f}")
    print(f"  VaR 95% (param)  -${m['var_95_param']:,.0f}   |  CVaR 95%        -${m['cvar_95']:,.0f}")
    print(f"  Sharpe   {m['sharpe']:>6.3f}    |  Sortino  {m['sortino']:>6.3f}    |  Beta  {m['beta']:>5.3f}")
    print(f"  Max DD   {m['max_drawdown_pct']:>6.2f}%")

    # Stress tests
    print(f"\n  {'─'*68}")
    print("  STRESS TESTS (choc instantané marché)")
    print(f"  {'─'*68}")
    for scenario, pnl in m['stress'].items():
        sign_s = '+' if pnl >= 0 else ''
        bar = '▼' * min(int(abs(pnl) / 500), 20) if pnl < 0 else '▲' * min(int(pnl / 500), 20)
        print(f"  Marché {scenario:<6}  P&L: {sign_s}${pnl:>8,.0f}   {bar}")

    print(f"\n  {'─'*68}")
    print("  [Entrée] Rafraîchir   [x + Entrée] Exporter Excel   [q + Entrée] Quitter")
    print("=" * 72)

# ── Export Excel ──────────────────────────────────────────────────────────────

def export_excel(snap: dict):
    if not snap:
        print("Pas de données à exporter.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Couleurs
    BLUE_HDR   = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
    GRAY_HDR   = PatternFill('solid', start_color='D6DCE4', end_color='D6DCE4')
    GREEN_CELL = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
    RED_CELL   = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
    WHITE_HDR  = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    BLACK_HDR  = Font(bold=True, color='000000', name='Arial', size=10)
    NORMAL     = Font(name='Arial', size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    m   = snap['metrics']
    df  = snap['positions']
    sa  = snap['strat_alloc']

    # ── Onglet 1 : Positions ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Positions'

    ws1['A1'] = f"RISK DASHBOARD — {m['timestamp']}"
    ws1['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4E79')
    ws1.merge_cells('A1:K1')
    ws1.row_dimensions[1].height = 25

    # Headers
    headers = list(df.columns)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.fill = BLUE_HDR
        cell.font = WHITE_HDR
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows(), 4):
        for col_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL
            cell.border = border
            cell.alignment = Alignment(horizontal='right' if col_idx > 3 else 'left')
            # Coloration P&L
            if col_idx == headers.index('Unreal. P&L ($)') + 1:
                cell.fill = GREEN_CELL if val >= 0 else RED_CELL
            if col_idx == headers.index('Return (%)') + 1:
                cell.fill = GREEN_CELL if val >= 0 else RED_CELL

    # Ligne totaux
    n_data = len(df)
    tot_row = 4 + n_data
    ws1.cell(row=tot_row, column=1, value='TOTAL').font = Font(bold=True, name='Arial')
    ws1.cell(row=tot_row, column=7, value=f'=SUM(G4:G{tot_row-1})').font = Font(bold=True, name='Arial')
    ws1.cell(row=tot_row, column=9, value=f'=SUM(I4:I{tot_row-1})').font = Font(bold=True, name='Arial')

    # Largeurs colonnes
    col_widths = [8, 14, 10, 8, 10, 10, 14, 14, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Onglet 2 : Risque ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Risque & VaR')

    def section(ws, row, title):
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11,
                                                             name='Arial', color='1F4E79')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 20
        return row + 1

    def kv(ws, row, key, value, fmt=None):
        kc = ws.cell(row=row, column=1, value=key)
        vc = ws.cell(row=row, column=2, value=value)
        kc.fill  = GRAY_HDR
        kc.font  = BLACK_HDR
        kc.border = border
        vc.font  = NORMAL
        vc.border = border
        vc.alignment = Alignment(horizontal='right')
        if fmt:
            vc.number_format = fmt
        return row + 1

    r = 1
    ws2['A1'] = f"Métriques de Risque — {m['timestamp']}"
    ws2['A1'].font = Font(bold=True, size=13, name='Arial', color='1F4E79')
    ws2.merge_cells('A1:D1')
    r = 3

    r = section(ws2, r, '📈 P&L & Exposition')
    r = kv(ws2, r, 'NAV ($)',            m['nav'],              '$#,##0.00')
    r = kv(ws2, r, 'Unreal. P&L ($)', m['total_upnl'],       '$#,##0.00')
    r = kv(ws2, r, 'Return (%)',         m['total_upnl_pct']/100, '0.00%')
    r = kv(ws2, r, 'Cash ($)',           m['cash'],             '$#,##0.00')
    r = kv(ws2, r, 'Gross Exposure ($)', m['gross_exposure'],   '$#,##0.00')
    r = kv(ws2, r, 'Net Exposure ($)',   m['net_exposure'],     '$#,##0.00')
    r += 1

    r = section(ws2, r, '📉 Value at Risk')
    r = kv(ws2, r, 'VaR 95% Historique',  -m['var_95_hist'],   '$#,##0.00')
    r = kv(ws2, r, 'VaR 99% Historique',  -m['var_99_hist'],   '$#,##0.00')
    r = kv(ws2, r, 'VaR 95% Paramétrique',-m['var_95_param'],  '$#,##0.00')
    r = kv(ws2, r, 'CVaR 95% (ES)',       -m['cvar_95'],       '$#,##0.00')
    r += 1

    r = section(ws2, r, '📊 Ratios Performance')
    r = kv(ws2, r, 'Sharpe Ratio',      m['sharpe'],            '0.000')
    r = kv(ws2, r, 'Sortino Ratio',     m['sortino'],           '0.000')
    r = kv(ws2, r, 'Beta vs SPY',       m['beta'],              '0.000')
    r = kv(ws2, r, 'Max Drawdown',      m['max_drawdown_pct']/100, '0.00%')
    r += 1

    r = section(ws2, r, '🔥 Stress Tests')
    for scenario, pnl in m['stress'].items():
        r = kv(ws2, r, f'Choc Marché {scenario}', pnl, '$#,##0.00')
        # Coloration stress
        cell = ws2.cell(row=r-1, column=2)
        cell.fill = RED_CELL if pnl < 0 else GREEN_CELL
    r += 1

    r = section(ws2, r, '🏦 Allocation par Stratégie')
    for strat, val in sorted(sa.items(), key=lambda x: -abs(x[1])):
        r = kv(ws2, r, strat, val, '$#,##0.00')

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16

    # ── Onglet 3 : Historique NAV ─────────────────────────────────────────────
    ws3 = wb.create_sheet('NAV History')
    ws3['A1'] = 'Index'
    ws3['B1'] = 'NAV ($)'
    for cell in [ws3['A1'], ws3['B1']]:
        cell.fill = BLUE_HDR
        cell.font = WHITE_HDR

    for i, nav_val in enumerate(pnl_history, 2):
        ws3.cell(row=i, column=1, value=i - 1)
        ws3.cell(row=i, column=2, value=round(nav_val, 2))

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 14

    # Sauvegarde
    fname = f"risk_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(fname)
    print(f"\n  ✅ Excel exporté : {fname}\n")
    return fname

# ── Boucle principale ─────────────────────────────────────────────────────────

def main():
    print("Connexion à IB TWS...")
    try:
        ib = connect()
    except Exception as e:
        print(f"Erreur connexion IB : {e}")
        alert_manager.notify_crash("risk_dashboard", f"Connexion IB impossible: {e}")
        sys.exit(1)

    ib.sleep(2)
    print("Connecté. Chargement du portefeuille...")

    snap = {}

    def refresh():
        nonlocal snap
        try:
            snap = build_snapshot(ib)
            print_dashboard(snap)
        except ConnectionError as e:
            log.error(f"IB connection lost: {e}")
            alert_manager.notify_crash('risk_dashboard', str(e))
        except Exception as e:
            log.error(f"Refresh error: {e}")
            alert_manager.notify_crash('risk_dashboard', str(e))

    # Premier chargement
    refresh()

    # Boucle input utilisateur
    while True:
        try:
            cmd = input().strip().lower()
        except (EOFError, KeyboardInterrupt):
            break

        if cmd == 'q':
            print("Déconnexion...")
            break
        elif cmd == 'x':
            fname = export_excel(snap)
            if fname:
                import shutil
                out = f"/mnt/user-data/outputs/{fname}"
                shutil.copy(fname, out)
                print(f"  Fichier disponible : {out}")
        else:
            refresh()

    ib.disconnect()


if __name__ == '__main__':
    main()
